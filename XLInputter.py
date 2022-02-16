#XLInputter App back-end:
#takes .csv-style formatted (text) data from a user-inputed file
#and fills it based off of a position (row, column) 
#given from user input (no built-in "framing")
#Author: Joshua Wilson Smith (https://www.github.com/jws8)
#Start Date: January 9, 2022
#Openpyxl is the only current python library needed

from openpyxl import Workbook, load_workbook
#Current working date: 
class XLInputter():
    def __init__(self):
        self.file_read = input("Enter file to be parsed: ") 
        self.wb_input = input("Into which Excel document? ")
        self.wb = load_workbook(self.wb_input) #create workbook 1/9/22
        self.row_size = input("row size: ") #"cutoff" 1/9/22
        self.row_pos = int(input("Enter row position start: ")) 
        self.col_pos = int(input("Enter column position start: "))
        self.saved_items = []
        self.ws = self.wb.active #create worksheet from workbook 1/9/22

    def save_file(self):
        self.wb.save(self.wb_input)

    def read_file(self):
        with open(self.file_read, "r") as f:
            for line in f:  #takes whole line 1/9/22
                self.saved_items = line.split(",") #split line into a list containing items seperated by ","'s. Then duplicates or "saves" this list into self.saved_items 1/31/22
                
    def parse_file(self):
        #col_pos, row_pos: these are so don't have to iterate over an enumerated (saved_items) list, for nothing.  Instead, iterate over
        #a saved_items list for item, and set item to the updated positions of
        #row and column. 1/31/22
        for item in (self.saved_items): 
            #too far to the right?!?! 1/31/22
            if self.col_pos > int(self.row_size):
                #if the col position (number) is greater than row size or the "cutoff", update the row position and reset the column position 1/31/22
                self.row_pos += 1
                self.col_pos = 1
            #fill the cell with the value associated with the updated row and col positions 1/31/22
            self.ws.cell(row = self.row_pos, column = self.col_pos).value = item
            #move to the right! Update the column position 1/31/22
            self.col_pos += 1
    #runs methods save_file, read_file, parse_file
    def run(self):
        print("""
        XLInputter App back-end:
        The executable runs an application XLInputter, that takes .csv-style formatted (text/string) data from a user-inputed file (example given in input_data_example.txt),
        and fills it based off of a position (row, column) given from user input (No built in "empty spaces" or "smart" framing.)
        Author: Joshua Wilson Smith (https://www.github.com/jws8)
        Date: January 31, 2022
        Openpyxl is the only current python library needed as a dependency. 
        Run the executable, or feel free to use any of the methods in XLInputter.py""")
        self.read_file()
        self.parse_file()
        self.save_file()

